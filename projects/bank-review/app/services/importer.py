"""
Orchestrates the full import pipeline:
  1. Parse export CSV
  2. Deduplicate against historical
  3. Classify (rules → Claude fallback)
  4. Return classified rows ready for review / append
"""
import pandas as pd
import numpy as np
from datetime import date
from typing import Optional

from app.services.deduplicator import find_new_transactions, normalize_amount
from app.services.classifier import TransactionClassifier, ClassificationResult, extract_real_date

USER_CATEGORIES = [
    "Nourriture", "Restaurant", "Vacances", "Transport", "Navigo", "Voiture",
    "Appartement", "Loyer", "Electricité", "Box", "Assurances", "Charges appartement",
    "Santé", "Pharmacie", "Crèche", "Loisirs", "Culture", "Cadeaux", "Dons",
    "Impôts", "Jean", "Amélie", "Paul", "Bébé", "Entrée mensuelle",
    "Entrée exceptionnelle", "Entrée", "Espèces", "Travaux", "Déménagement",
    "Notaire", "Taxe foncière", "Non catégorisé",
]


def load_historical(path: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name="Export", dtype=str)


def load_export_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", dtype=str)
    return df


async def run_import_pipeline(
    history_path: str,
    export_path: str,
    vacation_periods: list[tuple[date, date]] | None = None,
) -> list[dict]:
    """
    Returns a list of dicts, one per new transaction, with classification info.
    """
    df_history = load_historical(history_path)
    df_export = load_export_csv(export_path)

    # Deduplicate
    df_new = find_new_transactions(df_history, df_export)

    if df_new.empty:
        return []

    # Build classifier
    classifier = TransactionClassifier(vacation_periods=vacation_periods or [])
    classifier.set_history_sample(df_history)

    results: list[dict] = []
    pending_indices: list[int] = []

    # First pass: rules + vacation
    for i, row in df_new.iterrows():
        label = str(row.get("label", ""))
        date_op = str(row.get("dateOp", ""))
        bank_cat = str(row.get("category", ""))
        amount_raw = str(row.get("amount", "0")).replace(",", ".")
        try:
            amount = float(amount_raw)
        except ValueError:
            amount = 0.0

        result = classifier.classify_one(label, date_op, bank_cat)
        real_date = result.real_date or extract_real_date(label)

        entry = {
            "idx": int(i),
            "dateOp": date_op,
            "realDate": str(real_date) if real_date else None,
            "label": label,
            "amount": amount,
            "bankCategory": bank_cat,
            "category": result.category,
            "confidence": result.confidence,
            "method": result.method,
        }
        results.append(entry)
        if result.method == "pending":
            pending_indices.append(len(results) - 1)

    # Second pass: Claude for pending
    if pending_indices:
        pending_rows = [
            {
                "label": results[j]["label"],
                "bank_cat": results[j]["bankCategory"],
                "amount": results[j]["amount"],
            }
            for j in pending_indices
        ]
        claude_results = await classifier.classify_batch_with_claude(pending_rows, USER_CATEGORIES)
        for j, cr in zip(pending_indices, claude_results):
            results[j]["category"] = cr.category
            results[j]["confidence"] = cr.confidence
            results[j]["method"] = cr.method

    # Sort by dateOp descending
    results.sort(key=lambda r: r["dateOp"], reverse=True)
    return results


def append_to_historical(
    history_path: str,
    classified_rows: list[dict],
) -> int:
    """Append validated rows to the historical Excel file. Returns nb rows added."""
    df_history = load_historical(history_path)

    new_rows = []
    for row in classified_rows:
        date_op = row["dateOp"]
        mois = _to_month_label(date_op)
        new_rows.append({
            "Mois": mois,
            "Solde cpte": None,
            "DATE OPERATION": date_op,
            "DATE VALEUR": date_op,
            "LIBELLE": row["label"],
            "MONTANT": row["amount"],
            "DEVISE": "EUR",
            "Catégorie": row["category"],
            "Tag #1": row["bankCategory"],
            "Précision": None,
            "Précision sans wkd": None,
        })

    if not new_rows:
        return 0

    df_append = pd.DataFrame(new_rows)
    df_updated = pd.concat([df_history, df_append], ignore_index=True)

    # Rewrite Excel preserving the catégories sheet
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.load_workbook(history_path)
    ws = wb["Export"]

    # Clear existing data rows (keep header)
    ws.delete_rows(2, ws.max_row)
    for r_idx, df_row in enumerate(df_updated.itertuples(index=False), start=2):
        for c_idx, val in enumerate(df_row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val if val is not None and str(val) != "nan" else None)

    wb.save(history_path)
    return len(new_rows)


def _to_month_label(date_str: str) -> str:
    try:
        d = pd.to_datetime(date_str)
        return d.strftime("%m/%Y")
    except Exception:
        return ""
